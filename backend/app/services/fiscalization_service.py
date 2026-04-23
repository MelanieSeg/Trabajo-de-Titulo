from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import zipfile
from datetime import date, datetime, timedelta, timezone
from typing import Any, Literal

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import inspect as sqlalchemy_inspect
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.models import (
    ActivityLog,
    AuditTrailBlock,
    CertifiableReport,
    Company,
    ComplianceAssessment,
    ComplianceStandard,
    Facility,
    LegalRequirement,
    MLPrediction,
    MeasurementCertificate,
    MeterCalibration,
    MonthlyConsumption,
    ResourceAreaDistribution,
    ResourceMonthlyConsumption,
    ResourceType,
    SmartAlert,
)
from app.services.activity_service import log_activity
from app.utils.date_utils import month_label


def _serialize(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, list):
        return [_serialize(item) for item in value]
    if isinstance(value, dict):
        return {str(k): _serialize(v) for k, v in value.items()}
    return value


def _canonical_json_bytes(payload: Any) -> bytes:
    return json.dumps(_serialize(payload), ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def _sha256_hex(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _sign_payload(payload: Any) -> str:
    settings = get_settings()
    digest = hmac.new(
        key=settings.secret_key.encode("utf-8"),
        msg=_canonical_json_bytes(payload),
        digestmod=hashlib.sha256,
    ).hexdigest()
    return digest


def _to_row_dict(item: Any) -> dict[str, Any]:
    mapper = sqlalchemy_inspect(item.__class__)
    data: dict[str, Any] = {}
    for attr in mapper.column_attrs:
        key = attr.key
        col_name = attr.columns[0].name
        data[col_name] = _serialize(getattr(item, key))
    return data


def _latest_period(db: Session) -> tuple[int, int]:
    row = db.execute(
        select(MonthlyConsumption.year, MonthlyConsumption.month).order_by(
            MonthlyConsumption.year.desc(),
            MonthlyConsumption.month.desc(),
        )
    ).first()
    if not row:
        now = datetime.now(timezone.utc)
        return now.year, now.month
    return int(row.year), int(row.month)


def _period_totals(db: Session, year: int, month: int) -> dict[str, float]:
    rows = db.execute(
        select(
            MonthlyConsumption.electricity_kwh,
            MonthlyConsumption.water_m3,
            MonthlyConsumption.electricity_cost_usd,
            MonthlyConsumption.water_cost_usd,
            MonthlyConsumption.co2_avoided_ton,
        ).where(MonthlyConsumption.year == year, MonthlyConsumption.month == month)
    ).all()
    electricity = sum(float(r.electricity_kwh or 0) for r in rows)
    water = sum(float(r.water_m3 or 0) for r in rows)
    electricity_cost = sum(float(r.electricity_cost_usd or 0) for r in rows)
    water_cost = sum(float(r.water_cost_usd or 0) for r in rows)
    co2 = sum(float(r.co2_avoided_ton or 0) for r in rows)
    return {
        "electricity_kwh": round(electricity, 2),
        "water_m3": round(water, 2),
        "electricity_cost_usd": round(electricity_cost, 2),
        "water_cost_usd": round(water_cost, 2),
        "total_cost_usd": round(electricity_cost + water_cost, 2),
        "co2_avoided_ton": round(co2, 2),
    }


def _compare_limit(
    observed: float,
    operator: str,
    limit: float,
    warning_ratio: float,
) -> tuple[Literal["compliant", "warning", "breach"], Literal["low", "medium", "high", "critical"], float]:
    if limit <= 0:
        return "compliant", "low", 0.0

    status: Literal["compliant", "warning", "breach"]
    risk_level: Literal["low", "medium", "high", "critical"]

    if operator in {"<=", "<"}:
        ratio = observed / limit
        breached = observed > limit if operator == "<=" else observed >= limit
        warning = (not breached) and ratio >= warning_ratio
        risk_score = min(100.0, max(0.0, ratio * 100.0))
        if breached:
            status = "breach"
            risk_level = "critical" if ratio >= 1.2 else "high"
        elif warning:
            status = "warning"
            risk_level = "medium"
        else:
            status = "compliant"
            risk_level = "low"
    elif operator in {">=", ">"}:
        ratio = (limit / observed) if observed > 0 else 100.0
        breached = observed < limit if operator == ">=" else observed <= limit
        warning_threshold = 1 + (1 - warning_ratio)
        warning = (not breached) and ratio >= warning_threshold
        risk_score = min(100.0, max(0.0, ratio * 100.0))
        if breached:
            status = "breach"
            risk_level = "critical" if ratio >= 1.2 else "high"
        elif warning:
            status = "warning"
            risk_level = "medium"
        else:
            status = "compliant"
            risk_level = "low"
    else:
        diff = abs(observed - limit)
        risk_score = min(100.0, max(0.0, (diff / limit) * 100.0))
        if diff < 0.000001:
            status = "compliant"
            risk_level = "low"
        else:
            status = "warning"
            risk_level = "medium"

    return status, risk_level, round(risk_score, 2)


def _record_audit_block(
    db: Session,
    *,
    entity_type: str,
    entity_id: str,
    action: str,
    payload: dict[str, Any],
    actor: str = "system",
    metadata: dict[str, Any] | None = None,
) -> AuditTrailBlock:
    payload_hash = _sha256_hex(_canonical_json_bytes(payload))
    existing = db.scalar(
        select(AuditTrailBlock).where(
            AuditTrailBlock.entity_type == entity_type,
            AuditTrailBlock.entity_id == entity_id,
            AuditTrailBlock.action == action,
            AuditTrailBlock.payload_hash == payload_hash,
        )
    )
    if existing:
        return existing

    previous = db.scalar(select(AuditTrailBlock).order_by(AuditTrailBlock.id.desc()))
    previous_hash = previous.block_hash if previous else "GENESIS"
    block_hash = _sha256_hex(
        _canonical_json_bytes(
            {
                "entity_type": entity_type,
                "entity_id": entity_id,
                "action": action,
                "payload_hash": payload_hash,
                "previous_hash": previous_hash,
                "actor": actor,
            }
        )
    )
    signature = _sign_payload({"block_hash": block_hash, "payload_hash": payload_hash, "actor": actor})
    block = AuditTrailBlock(
        entity_type=entity_type,
        entity_id=entity_id,
        action=action,
        payload_hash=payload_hash,
        previous_hash=previous_hash,
        block_hash=block_hash,
        digital_signature=signature,
        actor=actor,
        metadata_json=metadata or {},
    )
    db.add(block)
    db.flush()
    return block


def list_standards(db: Session) -> list[ComplianceStandard]:
    return db.scalars(select(ComplianceStandard).order_by(ComplianceStandard.code.asc())).all()


def list_legal_requirements(db: Session, active_only: bool = True) -> list[LegalRequirement]:
    stmt = select(LegalRequirement).order_by(LegalRequirement.code.asc())
    if active_only:
        stmt = stmt.where(LegalRequirement.is_active.is_(True))
    return db.scalars(stmt).all()


def upsert_legal_requirement(db: Session, payload: dict[str, Any]) -> LegalRequirement:
    standard_id: int | None = None
    standard_code = payload.get("standard_code")
    if standard_code:
        standard = db.scalar(select(ComplianceStandard).where(ComplianceStandard.code == standard_code))
        if standard:
            standard_id = standard.id

    req = db.scalar(select(LegalRequirement).where(LegalRequirement.code == payload["code"]))
    if req:
        req.standard_id = standard_id
        req.title = payload["title"]
        req.utility = payload["utility"]
        req.metric_name = payload["metric_name"]
        req.limit_operator = payload["limit_operator"]
        req.limit_value = float(payload["limit_value"])
        req.limit_unit = payload["limit_unit"]
        req.warning_ratio = float(payload["warning_ratio"])
        req.severity_on_breach = payload["severity_on_breach"]
        req.jurisdiction = payload.get("jurisdiction")
        req.legal_reference = payload.get("legal_reference")
        req.is_active = True
    else:
        req = LegalRequirement(
            standard_id=standard_id,
            code=payload["code"],
            title=payload["title"],
            utility=payload["utility"],
            metric_name=payload["metric_name"],
            limit_operator=payload["limit_operator"],
            limit_value=float(payload["limit_value"]),
            limit_unit=payload["limit_unit"],
            warning_ratio=float(payload["warning_ratio"]),
            severity_on_breach=payload["severity_on_breach"],
            jurisdiction=payload.get("jurisdiction"),
            legal_reference=payload.get("legal_reference"),
            is_active=True,
        )
        db.add(req)
        db.flush()

    log_activity(
        db,
        activity_type="compliance_requirement",
        message=f"Requisito legal actualizado: {req.code}",
        metadata={"requirement_id": req.id, "metric_name": req.metric_name},
    )
    _record_audit_block(
        db,
        entity_type="legal_requirement",
        entity_id=str(req.id),
        action="upsert",
        payload={
            "code": req.code,
            "metric_name": req.metric_name,
            "limit_operator": req.limit_operator,
            "limit_value": req.limit_value,
            "warning_ratio": req.warning_ratio,
        },
        actor="system",
    )
    return req


def evaluate_compliance(
    db: Session,
    *,
    year: int | None = None,
    month: int | None = None,
    create_alerts: bool = True,
) -> dict[str, Any]:
    if year is None or month is None:
        period_year, period_month = _latest_period(db)
    else:
        period_year, period_month = year, month

    totals = _period_totals(db, period_year, period_month)
    requirements = list_legal_requirements(db, active_only=True)
    evaluations: list[dict[str, Any]] = []

    summary = {"compliant": 0, "warning": 0, "breach": 0}

    for requirement in requirements:
        observed = float(totals.get(requirement.metric_name, 0.0))
        status, risk_level, risk_score = _compare_limit(
            observed=observed,
            operator=requirement.limit_operator,
            limit=float(requirement.limit_value),
            warning_ratio=float(requirement.warning_ratio),
        )

        row = db.scalar(
            select(ComplianceAssessment).where(
                ComplianceAssessment.requirement_id == requirement.id,
                ComplianceAssessment.year == period_year,
                ComplianceAssessment.month == period_month,
            )
        )
        details = {
            "evaluated_metric": requirement.metric_name,
            "legal_reference": requirement.legal_reference,
            "jurisdiction": requirement.jurisdiction,
        }
        if row:
            row.observed_value = observed
            row.limit_value = float(requirement.limit_value)
            row.status = status
            row.risk_level = risk_level
            row.risk_score = risk_score
            row.details = details
        else:
            row = ComplianceAssessment(
                requirement_id=requirement.id,
                year=period_year,
                month=period_month,
                observed_value=observed,
                limit_value=float(requirement.limit_value),
                status=status,
                risk_level=risk_level,
                risk_score=risk_score,
                details=details,
            )
            db.add(row)
            db.flush()

        summary[status] += 1
        evaluations.append(
            {
                "requirement_id": requirement.id,
                "code": requirement.code,
                "title": requirement.title,
                "utility": requirement.utility,
                "metric_name": requirement.metric_name,
                "observed_value": round(observed, 2),
                "limit_operator": requirement.limit_operator,
                "limit_value": round(float(requirement.limit_value), 2),
                "unit": requirement.limit_unit,
                "status": status,
                "risk_level": risk_level,
                "risk_score": risk_score,
                "legal_reference": requirement.legal_reference,
            }
        )

        if not create_alerts or status == "compliant":
            continue

        alert_title = f"Riesgo de infracción normativa: {requirement.code}"
        existing_alert = db.scalar(
            select(SmartAlert).where(
                SmartAlert.is_resolved.is_(False),
                SmartAlert.title == alert_title,
                SmartAlert.year == period_year,
                SmartAlert.month == period_month,
            )
        )
        if existing_alert:
            continue

        severity = requirement.severity_on_breach if status == "breach" else "warning"
        description = (
            f"{requirement.title}: valor observado {observed:.2f} {requirement.limit_unit}, "
            f"límite {requirement.limit_operator} {float(requirement.limit_value):.2f} {requirement.limit_unit}."
        )
        alert = SmartAlert(
            severity=severity if severity in {"critical", "warning", "info"} else "warning",
            title=alert_title,
            description=description,
            utility=requirement.utility,
            year=period_year,
            month=period_month,
            is_resolved=False,
            extra_data={
                "type": "compliance_risk",
                "requirement_id": requirement.id,
                "requirement_code": requirement.code,
                "risk_level": risk_level,
                "risk_score": risk_score,
                "observed_value": observed,
                "limit_value": float(requirement.limit_value),
            },
        )
        db.add(alert)

    payload = {
        "year": period_year,
        "month": period_month,
        "month_label": month_label(period_year, period_month),
        "summary": summary,
        "evaluations": evaluations,
    }

    log_activity(
        db,
        activity_type="compliance_evaluation",
        message=f"Evaluación normativa {month_label(period_year, period_month)}",
        metadata=payload["summary"],
    )
    _record_audit_block(
        db,
        entity_type="compliance_assessment",
        entity_id=f"{period_year}-{period_month:02d}",
        action="evaluate",
        payload=payload,
        actor="system",
    )
    return payload


def _build_pdf_bytes(payload: dict[str, Any], metadata: dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 45

    def write_line(text: str, step: int = 14) -> None:
        nonlocal y
        if y < 40:
            pdf.showPage()
            y = height - 45
        pdf.drawString(40, y, text)
        y -= step

    pdf.setFont("Helvetica-Bold", 13)
    write_line("Reporte Formal de Fiscalizacion", step=20)
    pdf.setFont("Helvetica", 10)
    write_line(f"Periodo: {payload['month_label']} ({payload['year']}-{payload['month']:02d})")
    write_line(f"Generado UTC: {metadata['generated_at'].isoformat()}")
    write_line(f"Codigo reporte: {metadata['report_code']}")
    write_line(f"SHA256: {metadata['sha256_hash']}")
    write_line(f"Firma digital: {metadata['digital_signature']}")
    write_line("")
    write_line("Resumen de consumo")
    totals = payload["totals"]
    write_line(f"- Electricidad: {totals['electricity_kwh']:.2f} kWh")
    write_line(f"- Agua: {totals['water_m3']:.2f} m3")
    write_line(f"- Costo total: {totals['total_cost_usd']:.2f} USD")
    write_line(f"- CO2 evitado: {totals['co2_avoided_ton']:.2f} Ton")
    write_line("")
    write_line("Cumplimiento normativo")
    write_line(
        "- Compliant: {compliant} | Warning: {warning} | Breach: {breach}".format(
            **payload["compliance"]["summary"]
        )
    )
    for item in payload["compliance"]["evaluations"]:
        write_line(
            f"* {item['code']}: {item['status']} "
            f"(obs {item['observed_value']:.2f} {item['unit']} "
            f"{item['limit_operator']} {item['limit_value']:.2f} {item['unit']})"
        )
    write_line("")
    write_line("Calibraciones")
    write_line(
        f"- Vigentes: {payload['calibrations']['valid']} | "
        f"Por vencer: {payload['calibrations']['expiring']} | "
        f"Expiradas: {payload['calibrations']['expired']}"
    )

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer.read()


def _build_excel_bytes(payload: dict[str, Any], metadata: dict[str, Any]) -> bytes:
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "Resumen"
    ws_summary.append(["Campo", "Valor"])
    ws_summary.append(["Periodo", payload["month_label"]])
    ws_summary.append(["Year", payload["year"]])
    ws_summary.append(["Month", payload["month"]])
    ws_summary.append(["Generado UTC", metadata["generated_at"].isoformat()])
    ws_summary.append(["Codigo reporte", metadata["report_code"]])
    ws_summary.append(["SHA256", metadata["sha256_hash"]])
    ws_summary.append(["Firma digital", metadata["digital_signature"]])
    ws_summary.append(["Electricidad (kWh)", payload["totals"]["electricity_kwh"]])
    ws_summary.append(["Agua (m3)", payload["totals"]["water_m3"]])
    ws_summary.append(["Costo total (USD)", payload["totals"]["total_cost_usd"]])
    ws_summary.append(["CO2 evitado (Ton)", payload["totals"]["co2_avoided_ton"]])

    ws_comp = workbook.create_sheet("Cumplimiento")
    ws_comp.append(
        [
            "Codigo",
            "Titulo",
            "Metrica",
            "Observado",
            "Operador",
            "Limite",
            "Unidad",
            "Estado",
            "Riesgo",
            "Puntaje",
            "Referencia legal",
        ]
    )
    for item in payload["compliance"]["evaluations"]:
        ws_comp.append(
            [
                item["code"],
                item["title"],
                item["metric_name"],
                item["observed_value"],
                item["limit_operator"],
                item["limit_value"],
                item["unit"],
                item["status"],
                item["risk_level"],
                item["risk_score"],
                item["legal_reference"] or "",
            ]
        )

    ws_cal = workbook.create_sheet("Calibraciones")
    ws_cal.append(["Estado", "Cantidad"])
    ws_cal.append(["valid", payload["calibrations"]["valid"]])
    ws_cal.append(["expiring", payload["calibrations"]["expiring"]])
    ws_cal.append(["expired", payload["calibrations"]["expired"]])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def _report_payload(db: Session, year: int, month: int) -> dict[str, Any]:
    totals = _period_totals(db, year, month)
    compliance = evaluate_compliance(db, year=year, month=month, create_alerts=True)
    calibrations = list_calibrations(db, limit=100)
    counters = {"valid": 0, "expiring": 0, "expired": 0}
    for item in calibrations:
        counters[item["status"]] += 1

    return {
        "year": year,
        "month": month,
        "month_label": month_label(year, month),
        "totals": totals,
        "compliance": compliance,
        "calibrations": counters,
    }


def generate_certifiable_report(
    db: Session,
    *,
    report_format: Literal["pdf", "xlsx"],
    year: int | None = None,
    month: int | None = None,
    generated_by: str = "system",
) -> dict[str, Any]:
    if year is None or month is None:
        period_year, period_month = _latest_period(db)
    else:
        period_year, period_month = year, month

    payload = _report_payload(db, period_year, period_month)
    generated_at = datetime.now(timezone.utc)
    report_code = f"RPT-{period_year}{period_month:02d}-{report_format.upper()}-{generated_at.strftime('%Y%m%d%H%M%S')}"

    metadata = {
        "report_code": report_code,
        "report_name": f"Reporte Fiscalizacion {month_label(period_year, period_month)}",
        "report_format": report_format,
        "year": period_year,
        "month": period_month,
        "generated_at": generated_at,
    }

    if report_format == "pdf":
        raw_content = _build_pdf_bytes(payload, {**metadata, "sha256_hash": "", "digital_signature": ""})
        mime = "application/pdf"
        ext = "pdf"
    else:
        raw_content = _build_excel_bytes(payload, {**metadata, "sha256_hash": "", "digital_signature": ""})
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"

    file_hash = _sha256_hex(raw_content)
    signature = _sign_payload(
        {
            "report_code": report_code,
            "sha256_hash": file_hash,
            "generated_at": generated_at.isoformat(),
            "format": report_format,
        }
    )
    verification_token = _sign_payload(
        {
            "report_code": report_code,
            "sha256_hash": file_hash,
            "digital_signature": signature,
            "generated_at": generated_at.isoformat(),
        }
    )

    report_row = CertifiableReport(
        report_code=report_code,
        report_name=metadata["report_name"],
        report_format=report_format,
        year=period_year,
        month=period_month,
        sha256_hash=file_hash,
        digital_signature=signature,
        timestamp_utc=generated_at,
        generated_by=generated_by,
        metadata_json={
            "verification_token": verification_token,
            "summary": payload["compliance"]["summary"],
            "totals": payload["totals"],
        },
    )
    db.add(report_row)
    db.flush()

    _record_audit_block(
        db,
        entity_type="certifiable_report",
        entity_id=str(report_row.id),
        action="generate",
        payload={
            "report_code": report_row.report_code,
            "sha256_hash": report_row.sha256_hash,
            "report_format": report_row.report_format,
            "year": report_row.year,
            "month": report_row.month,
        },
        actor=generated_by,
    )
    log_activity(
        db,
        activity_type="certifiable_report",
        message=f"Reporte certificable generado ({report_row.report_format.upper()}): {report_row.report_code}",
        metadata={
            "report_id": report_row.id,
            "sha256_hash": report_row.sha256_hash,
        },
    )

    filename = f"{report_row.report_code}.{ext}"
    return {
        "bytes": raw_content,
        "mime": mime,
        "filename": filename,
        "metadata": {
            "report_code": report_code,
            "report_name": report_row.report_name,
            "report_format": report_format,
            "year": period_year,
            "month": period_month,
            "generated_at": generated_at,
            "sha256_hash": file_hash,
            "digital_signature": signature,
            "verification_token": verification_token,
        },
    }


def list_report_history(db: Session, limit: int = 50) -> list[dict[str, Any]]:
    rows = db.scalars(select(CertifiableReport).order_by(CertifiableReport.created_at.desc()).limit(limit)).all()
    payload: list[dict[str, Any]] = []
    for row in rows:
        payload.append(
            {
                "report_code": row.report_code,
                "report_name": row.report_name,
                "report_format": row.report_format,
                "year": row.year,
                "month": row.month,
                "generated_at": row.timestamp_utc,
                "sha256_hash": row.sha256_hash,
                "digital_signature": row.digital_signature,
                "verification_token": (row.metadata_json or {}).get("verification_token", ""),
            }
        )
    return payload


def _calibration_status(valid_until: datetime, now: datetime) -> Literal["valid", "expiring", "expired"]:
    if valid_until <= now:
        return "expired"
    if valid_until <= now + timedelta(days=30):
        return "expiring"
    return "valid"


def create_calibration(db: Session, payload: dict[str, Any]) -> dict[str, Any]:
    now = datetime.now(timezone.utc)
    calibrated_at = payload["calibrated_at"]
    valid_until = payload["valid_until"]
    status = _calibration_status(valid_until, now)

    calibration = MeterCalibration(
        meter_code=payload["meter_code"],
        meter_name=payload["meter_name"],
        facility_name=payload.get("facility_name"),
        utility=payload["utility"],
        calibrated_at=calibrated_at,
        valid_until=valid_until,
        performed_by=payload["performed_by"],
        status=status,
        metadata_json={"notes": payload.get("notes")},
    )
    db.add(calibration)
    db.flush()

    cert_number = f"CERT-{calibrated_at.strftime('%Y%m%d')}-{calibration.id:06d}"
    cert_payload = {
        "certificate_number": cert_number,
        "calibration_id": calibration.id,
        "meter_code": calibration.meter_code,
        "meter_name": calibration.meter_name,
        "facility_name": calibration.facility_name,
        "utility": calibration.utility,
        "calibrated_at": calibration.calibrated_at,
        "valid_until": calibration.valid_until,
        "performed_by": calibration.performed_by,
    }
    cert_hash = _sha256_hex(_canonical_json_bytes(cert_payload))
    cert_signature = _sign_payload({"certificate_number": cert_number, "sha256_hash": cert_hash})

    certificate = MeasurementCertificate(
        calibration_id=calibration.id,
        certificate_number=cert_number,
        issued_at=now,
        expires_at=valid_until,
        issuer=payload["performed_by"],
        sha256_hash=cert_hash,
        digital_signature=cert_signature,
        payload=cert_payload,
    )
    db.add(certificate)
    db.flush()

    _record_audit_block(
        db,
        entity_type="calibration",
        entity_id=str(calibration.id),
        action="create",
        payload=cert_payload,
        actor=payload["performed_by"],
        metadata={"certificate_number": cert_number},
    )
    log_activity(
        db,
        activity_type="calibration",
        message=f"Calibracion registrada para medidor {calibration.meter_code}",
        metadata={"calibration_id": calibration.id, "certificate_number": cert_number},
    )

    return {
        "id": calibration.id,
        "meter_code": calibration.meter_code,
        "meter_name": calibration.meter_name,
        "facility_name": calibration.facility_name,
        "utility": calibration.utility,
        "calibrated_at": calibration.calibrated_at,
        "valid_until": calibration.valid_until,
        "performed_by": calibration.performed_by,
        "status": calibration.status,
        "certificate_number": cert_number,
    }


def list_calibrations(db: Session, limit: int = 100) -> list[dict[str, Any]]:
    now = datetime.now(timezone.utc)
    rows = db.scalars(select(MeterCalibration).order_by(MeterCalibration.valid_until.asc()).limit(limit)).all()
    if not rows:
        return []

    certs = db.scalars(select(MeasurementCertificate).where(MeasurementCertificate.calibration_id.in_([r.id for r in rows]))).all()
    cert_map = {c.calibration_id: c for c in certs}

    payload: list[dict[str, Any]] = []
    for row in rows:
        computed = _calibration_status(row.valid_until, now)
        if row.status != computed:
            row.status = computed
        cert = cert_map.get(row.id)
        payload.append(
            {
                "id": row.id,
                "meter_code": row.meter_code,
                "meter_name": row.meter_name,
                "facility_name": row.facility_name,
                "utility": row.utility,
                "calibrated_at": row.calibrated_at,
                "valid_until": row.valid_until,
                "performed_by": row.performed_by,
                "status": row.status,
                "certificate_number": cert.certificate_number if cert else None,
            }
        )
    return payload


def get_calibration_certificate(db: Session, calibration_id: int) -> dict[str, Any]:
    cert = db.scalar(
        select(MeasurementCertificate).where(MeasurementCertificate.calibration_id == calibration_id)
    )
    if not cert:
        raise ValueError("No existe certificado para la calibracion indicada")
    return {
        "calibration_id": calibration_id,
        "certificate_number": cert.certificate_number,
        "issued_at": cert.issued_at,
        "expires_at": cert.expires_at,
        "issuer": cert.issuer,
        "sha256_hash": cert.sha256_hash,
        "digital_signature": cert.digital_signature,
        "payload": cert.payload,
    }


def verify_audit_chain(db: Session) -> dict[str, Any]:
    rows = db.scalars(select(AuditTrailBlock).order_by(AuditTrailBlock.id.asc())).all()
    previous_hash = "GENESIS"
    for row in rows:
        expected_hash = _sha256_hex(
            _canonical_json_bytes(
                {
                    "entity_type": row.entity_type,
                    "entity_id": row.entity_id,
                    "action": row.action,
                    "payload_hash": row.payload_hash,
                    "previous_hash": row.previous_hash,
                    "actor": row.actor,
                }
            )
        )
        if row.previous_hash != previous_hash:
            return {
                "valid": False,
                "total_blocks": len(rows),
                "broken_block_id": row.id,
                "checked_at": datetime.now(timezone.utc),
                "message": "La cadena de auditoria tiene un enlace roto.",
            }
        if expected_hash != row.block_hash:
            return {
                "valid": False,
                "total_blocks": len(rows),
                "broken_block_id": row.id,
                "checked_at": datetime.now(timezone.utc),
                "message": "La huella del bloque no coincide con el contenido auditado.",
            }
        previous_hash = row.block_hash

    return {
        "valid": True,
        "total_blocks": len(rows),
        "broken_block_id": None,
        "checked_at": datetime.now(timezone.utc),
        "message": "Cadena de auditoria integra.",
    }


def _rows_to_csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    output = io.StringIO()
    if not rows:
        output.write("")
        return output.getvalue().encode("utf-8")
    headers = list(rows[0].keys())
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow({k: _serialize(v) for k, v in row.items()})
    return output.getvalue().encode("utf-8")


def export_activity_logs(db: Session, export_format: Literal["csv", "json"]) -> bytes:
    logs = db.scalars(select(ActivityLog).order_by(ActivityLog.created_at.asc())).all()
    payload = [_to_row_dict(item) for item in logs]
    if export_format == "json":
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    return _rows_to_csv_bytes(payload)


def export_audit_chain(db: Session, export_format: Literal["csv", "json"]) -> bytes:
    blocks = db.scalars(select(AuditTrailBlock).order_by(AuditTrailBlock.id.asc())).all()
    payload = [_to_row_dict(item) for item in blocks]
    if export_format == "json":
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    return _rows_to_csv_bytes(payload)


def _collect_raw_datasets(db: Session) -> dict[str, list[dict[str, Any]]]:
    datasets: dict[str, list[dict[str, Any]]] = {
        "companies": [_to_row_dict(item) for item in db.scalars(select(Company).order_by(Company.id.asc())).all()],
        "facilities": [_to_row_dict(item) for item in db.scalars(select(Facility).order_by(Facility.id.asc())).all()],
        "monthly_consumptions": [
            _to_row_dict(item)
            for item in db.scalars(select(MonthlyConsumption).order_by(MonthlyConsumption.year.asc(), MonthlyConsumption.month.asc())).all()
        ],
        "ml_predictions": [_to_row_dict(item) for item in db.scalars(select(MLPrediction).order_by(MLPrediction.id.asc())).all()],
        "smart_alerts": [_to_row_dict(item) for item in db.scalars(select(SmartAlert).order_by(SmartAlert.id.asc())).all()],
        "activity_logs": [_to_row_dict(item) for item in db.scalars(select(ActivityLog).order_by(ActivityLog.id.asc())).all()],
        "compliance_standards": [_to_row_dict(item) for item in db.scalars(select(ComplianceStandard).order_by(ComplianceStandard.id.asc())).all()],
        "legal_requirements": [_to_row_dict(item) for item in db.scalars(select(LegalRequirement).order_by(LegalRequirement.id.asc())).all()],
        "compliance_assessments": [
            _to_row_dict(item) for item in db.scalars(select(ComplianceAssessment).order_by(ComplianceAssessment.id.asc())).all()
        ],
        "meter_calibrations": [_to_row_dict(item) for item in db.scalars(select(MeterCalibration).order_by(MeterCalibration.id.asc())).all()],
        "measurement_certificates": [
            _to_row_dict(item) for item in db.scalars(select(MeasurementCertificate).order_by(MeasurementCertificate.id.asc())).all()
        ],
        "certifiable_reports": [_to_row_dict(item) for item in db.scalars(select(CertifiableReport).order_by(CertifiableReport.id.asc())).all()],
        "audit_trail_blocks": [_to_row_dict(item) for item in db.scalars(select(AuditTrailBlock).order_by(AuditTrailBlock.id.asc())).all()],
        "resource_types": [_to_row_dict(item) for item in db.scalars(select(ResourceType).order_by(ResourceType.id.asc())).all()],
        "resource_monthly_consumptions": [
            _to_row_dict(item)
            for item in db.scalars(select(ResourceMonthlyConsumption).order_by(ResourceMonthlyConsumption.id.asc())).all()
        ],
        "resource_area_distributions": [
            _to_row_dict(item)
            for item in db.scalars(select(ResourceAreaDistribution).order_by(ResourceAreaDistribution.id.asc())).all()
        ],
    }
    return datasets


def _build_export_manifest(datasets: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    sections: list[dict[str, Any]] = []
    for name, rows in datasets.items():
        section_hash = _sha256_hex(_canonical_json_bytes(rows))
        sections.append({"table": name, "rows": len(rows), "sha256_hash": section_hash})

    generated_at = datetime.now(timezone.utc)
    overall_hash = _sha256_hex(_canonical_json_bytes(sections))
    signature = _sign_payload({"generated_at": generated_at.isoformat(), "overall_hash": overall_hash})
    return {
        "generated_at": generated_at.isoformat(),
        "overall_hash": overall_hash,
        "digital_signature": signature,
        "sections": sections,
    }


def export_raw_data_json(db: Session) -> bytes:
    datasets = _collect_raw_datasets(db)
    manifest = _build_export_manifest(datasets)
    payload = {"manifest": manifest, "data": datasets}
    _record_audit_block(
        db,
        entity_type="raw_export",
        entity_id=f"json-{manifest['generated_at']}",
        action="export",
        payload={"format": "json", "overall_hash": manifest["overall_hash"]},
        actor="system",
    )
    log_activity(
        db,
        activity_type="raw_export",
        message="Exportacion completa de datos crudos en JSON",
        metadata={"overall_hash": manifest["overall_hash"]},
    )
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def export_raw_data_zip(db: Session) -> bytes:
    datasets = _collect_raw_datasets(db)
    manifest = _build_export_manifest(datasets)

    output = io.BytesIO()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, rows in datasets.items():
            zf.writestr(f"{name}.csv", _rows_to_csv_bytes(rows))
            zf.writestr(f"{name}.json", json.dumps(rows, ensure_ascii=False, indent=2))
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    _record_audit_block(
        db,
        entity_type="raw_export",
        entity_id=f"zip-{manifest['generated_at']}",
        action="export",
        payload={"format": "zip", "overall_hash": manifest["overall_hash"]},
        actor="system",
    )
    log_activity(
        db,
        activity_type="raw_export",
        message="Exportacion completa de datos crudos en ZIP auditable",
        metadata={"overall_hash": manifest["overall_hash"]},
    )
    output.seek(0)
    return output.read()
